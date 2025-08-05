import pandas as pd
import numpy as np
import random
from datetime import datetime, timedelta
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils.dataframe import dataframe_to_rows

def generate_qingyu_survival_data(num_records=500):
    """
    生成青鱼养殖存活率综合环境分析数据
    
    Args:
        num_records: 生成记录数量，默认500条
    
    Returns:
        DataFrame: 包含所有字段的数据表
    """
    
    # 设置随机种子确保可重复性
    np.random.seed(42)
    random.seed(42)
    
    # 基础配置
    pool_prefixes = ['QY', 'CY', 'BY', 'DY', 'EY']
    death_reasons = ['pH值异常', '疾病', '操作损伤', '天敌', '水质异常', '缺氧', '温度异常']
    managers = ['张三', '李四', '王五', '赵六', '钱七', '孙八', '周九', '吴十']
    
    data = []
    
    for i in range(num_records):
        # 生成基础数据
        pool_id = f"{random.choice(pool_prefixes)}{str(random.randint(1, 99)).zfill(2)}"
        
        # 投放时间：2023年1月1日到2024年12月31日之间随机
        start_date = datetime(2023, 1, 1) + timedelta(days=random.randint(0, 730))
        release_date = start_date.strftime('%Y-%m-%d')
        
        # 投放数量：3000-8000尾
        release_count = random.randint(3000, 8000)
        
        # 养殖周期：60-180天
        culture_period = random.randint(60, 180)
        
        # 检测时间：投放后随机1-30天
        check_days = random.randint(1, 30)
        check_date = start_date + timedelta(days=check_days)
        check_date_str = check_date.strftime('%Y-%m-%d')
        
        # 环境参数（基于最优值生成合理范围）
        water_temp = round(random.uniform(20.0, 28.0), 1)  # 水温20-28℃
        dissolved_oxygen = round(random.uniform(5.0, 9.0), 1)  # 溶氧5-9mg/L
        ph_value = round(random.uniform(6.5, 8.0), 1)  # pH 6.5-8.0
        ammonia_nitrogen = round(random.uniform(0.05, 0.3), 2)  # 氨氮0.05-0.3mg/L
        
        # 根据环境参数计算存活率
        # 环境修正因子计算
        do_deviation = abs(dissolved_oxygen - 6.0) / 6.0
        temp_deviation = abs(water_temp - 24.0) / 24.0
        ph_deviation = abs(ph_value - 7.0) / 7.0
        nh3_deviation = abs(ammonia_nitrogen - 0.1) / 0.1
        
        env_correction_factor = 1 - (0.35 * do_deviation + 0.30 * temp_deviation + 
                                    0.20 * ph_deviation + 0.15 * nh3_deviation)
        env_correction_factor = max(0.5, min(1.0, env_correction_factor))  # 限制在0.5-1.0之间
        
        # 基础存活率：85%-98%
        base_survival_rate = random.uniform(0.85, 0.98)
        
        # 综合存活率
        comprehensive_survival_rate = base_survival_rate * env_correction_factor
        
        # 存活数量
        survival_count = int(release_count * comprehensive_survival_rate)
        
        # 死亡原因（根据环境参数判断）
        if ph_value < 6.8 or ph_value > 7.5:
            death_reason = 'pH值异常'
        elif dissolved_oxygen < 5.5:
            death_reason = '缺氧'
        elif water_temp < 22 or water_temp > 26:
            death_reason = '温度异常'
        elif ammonia_nitrogen > 0.2:
            death_reason = '水质异常'
        else:
            death_reason = random.choice(death_reasons)
        
        # 构建记录
        record = {
            '养殖池编号': pool_id,
            '投放时间': release_date,
            '投放数量（尾）': release_count,
            '存活数量（尾）': survival_count,
            '死亡原因': death_reason,
            '养殖周期（天）': culture_period,
            '检测时间': check_date_str,
            '负责人': random.choice(managers),
            '水温（℃）': water_temp,
            '溶氧（mg/L）': dissolved_oxygen,
            'pH': ph_value,
            '氨氮（mg/L）': ammonia_nitrogen,
            '基础存活率（%）': round(base_survival_rate * 100, 1),
            '综合存活率（%）': round(comprehensive_survival_rate * 100, 1)
        }
        
        data.append(record)
    
    return pd.DataFrame(data)

def create_excel_with_formatting(df, filename='青鱼养殖存活率综合环境分析数据.xlsx'):
    """
    创建格式化的Excel文件
    
    Args:
        df: DataFrame数据
        filename: 输出文件名
    """
    
    # 创建Excel写入器
    with pd.ExcelWriter(filename, engine='openpyxl') as writer:
        df.to_excel(writer, sheet_name='青鱼存活率数据', index=False)
        
        # 获取工作表
        worksheet = writer.sheets['青鱼存活率数据']
        
        # 设置列宽
        column_widths = {
            'A': 12,  # 养殖池编号
            'B': 12,  # 投放时间
            'C': 15,  # 投放数量（尾）
            'D': 15,  # 存活数量（尾）
            'E': 20,  # 死亡原因
            'F': 15,  # 养殖周期（天）
            'G': 12,  # 检测时间
            'H': 10,  # 负责人
            'I': 12,  # 水温（℃）
            'J': 15,  # 溶氧（mg/L）
            'K': 8,   # pH
            'L': 15,  # 氨氮（mg/L）
            'M': 18,  # 基础存活率（%）
            'N': 18   # 综合存活率（%）
        }
        
        for col, width in column_widths.items():
            worksheet.column_dimensions[col].width = width
        
        # 设置标题行格式
        header_font = Font(bold=True, color='FFFFFF')
        header_fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
        header_alignment = Alignment(horizontal='center', vertical='center')
        
        for cell in worksheet[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
        
        # 设置数据行格式
        data_alignment = Alignment(horizontal='center', vertical='center')
        for row in worksheet.iter_rows(min_row=2, max_row=len(df)+1):
            for cell in row:
                cell.alignment = data_alignment
        
        # 添加边框
        from openpyxl.styles import Border, Side
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        for row in worksheet.iter_rows(min_row=1, max_row=len(df)+1):
            for cell in row:
                cell.border = thin_border

def main():
    """主函数"""
    print("开始生成青鱼养殖存活率数据...")
    
    # 生成数据
    df = generate_qingyu_survival_data(500)
    
    # 创建Excel文件
    filename = '青鱼养殖存活率综合环境分析数据.xlsx'
    create_excel_with_formatting(df, filename)
    
    print(f"数据生成完成！共生成 {len(df)} 条记录")
    print(f"文件保存为: {filename}")
    
    # 显示数据统计信息
    print("\n数据统计信息:")
    print(f"平均基础存活率: {df['基础存活率（%）'].mean():.1f}%")
    print(f"平均综合存活率: {df['综合存活率（%）'].mean():.1f}%")
    print(f"平均水温: {df['水温（℃）'].mean():.1f}℃")
    print(f"平均溶氧: {df['溶氧（mg/L）'].mean():.1f}mg/L")
    print(f"平均pH值: {df['pH'].mean():.1f}")
    print(f"平均氨氮: {df['氨氮（mg/L）'].mean():.2f}mg/L")
    
    # 显示前5条数据预览
    print("\n前5条数据预览:")
    print(df.head().to_string(index=False))

if __name__ == "__main__":
    main() 